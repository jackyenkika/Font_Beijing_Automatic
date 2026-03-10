#auto_workflow.py
import os
import time
import requests
import app_context
from openpyxl import load_workbook
from util import log, get_system_chrome_path
from datetime import datetime
from urllib.parse import urlparse
from playwright.sync_api import sync_playwright

# ===============================
# 基本設定
# ===============================
LOGIN_URL = "https://banshi.beijing.gov.cn/pubservice/user/yztLogin?backUrl=/proAccept/urlCheck?baseUrl=aHR0cDovL2JqdC5iZWlqaW5nLmdvdi5jbi9yZW56aGVuZy9vcGVuL2xvZ2luL2dvVXNlckxvZ2luP2NsaWVudF9pZD0yMjc3JnJlZGlyZWN0X3VyaT1odHRwOi8vY29weXJpZ2h0LmJqeHdjYmouZ292LmNuL3ovaW5kZXgmcmVzcG9uc2VfdHlwZT1jb2RlJnNjb3BlPXVzZXJfaW5mbyZzdGF0ZT07aHR0cDovL3l6dC5iZWlqaW5nLmdvdi5jbi9hbS9vYXV0aDIvYXV0aG9yaXplP3NlcnZpY2U9Ymp6d1NlcnZpY2UmcmVzcG9uc2VfdHlwZT1jb2RlJmNsaWVudF9pZD0wMDAwMjExMDBfMDkmc2NvcGU9Y24rdWlkK2lkQ2FyZE51bWJlcitleHRQcm9wZXJ0aWVzK3Jlc2VydmUzJnJlZGlyZWN0X3VyaT1odHRwJTNBLy9jb3B5cmlnaHQuYmp4d2Niai5nb3YuY24vWVpUQXV0aCUzRm1hdHRlckNvZGUlM0QwNzAwMTQyMDAwMTEwMDAwMDAwMDAwMDAwMDIwMDQ0MTAw&serverType=1^2^3^4^5^6^9&isNewEvent=1"
ADD_NEW_URL = "http://copyright.bjxwcbj.gov.cn/z/voluntary/add_new?type=0"

# ===============================
# Utils
# ===============================
def wait_overlay_gone(page, timeout=10000):
    page.wait_for_selector(
        ".modal-backdrop, .loading, .layui-layer-shade",
        state="hidden",
        timeout=timeout
    )

def close_datepicker(page):
    page.keyboard.press("Escape")
    page.mouse.click(10, 10)
    page.evaluate("""
        document.querySelectorAll('.layui-laydate').forEach(el => el.remove());
    """)

def click_item(page, selector, text=None, delay=0):
    btn = page.locator(selector, has_text=text) if text else page.locator(selector)
    btn.scroll_into_view_if_needed()
    btn.wait_for(state="visible", timeout=5000)
    btn.click(delay=delay)

def prevent_pdf_new_tab(page):
    page.evaluate("""
    () => {
        const original_open = window.open;
        window.open = (url, name, features) => {
            console.log('阻止 window.open', url);
            // 如果是 PDF URL，就不打開
            if(url.includes('file.bjxwcbj.gov.cn')) return null;
            return original_open ? original_open(url, name, features) : null;
        };
        // 避免 a[target=_blank] 點擊打開新頁
        document.querySelectorAll('a[target="_blank"]').forEach(a=>{
            a.addEventListener('click', e=>{
                if(a.href.includes('file.bjxwcbj.gov.cn')) e.preventDefault();
            });
        });
    }
    """)

def upload_file(page, upload_section_selector, file_path, page_num=1, file_type="pdf"):
    section = page.locator(upload_section_selector)
    # 選份數
    section.locator("select.wt-batch-upload-select").select_option(value="1")
    # 選類型
    section.locator("select.wt-batch-upload-fileType").select_option(value=file_type)
    
    # 填頁數
    page_input = section.locator("input.wt-batch-upload-filePageNum:not([disabled]):not([readonly])").first
    page_input.fill(str(page_num))

    prevent_pdf_new_tab(page)
    with page.expect_file_chooser() as fc:
        section.locator(".wt-batch-upload-item .wt-btn-batch-upload").click()
    fc.value.set_files(file_path)

    page.evaluate("""
    () => {
      const input = document.querySelector(
        '.beforeNamediv input.webuploader-element-invisible[type=file]'
      );
      input.dispatchEvent(new Event('change', { bubbles: true }));
    }
    """)
    page.evaluate("""
    () => {
        const uploader = window.WebUploader?.instances?.[0];
        if (uploader) uploader.upload();
    }
    """)
    page.wait_for_function("""
    () => {
        const url = document.querySelector("input[name$='.file_url']")?.value;
        const pageNum = document.querySelector("input[name$='.file_page']")?.value;
        const kind = document.querySelector("input[name$='.file_kind']")?.value;
        return url && pageNum && kind;
    }
    """, timeout=30000)

def fill_person_block(page, container_selector, radio_name, sign_select_name):
    page.wait_for_function(
        f"""
        () => {{
            const divs = Array.from(document.querySelectorAll('{container_selector}'));
            return divs.some(d => d.offsetParent !== null && d.style.display !== 'none');
        }}
        """,
        timeout=10000
    )

    page.evaluate(
        f"""
        () => {{
            const divs = Array.from(document.querySelectorAll('{container_selector}'));
            const visibleDiv = divs.find(d => d.offsetParent !== null && d.style.display !== 'none');
            if (!visibleDiv) return;

            const radio = visibleDiv.querySelector(
                'input[type="radio"][name="{radio_name}"]'
            );
            if (radio) {{
                radio.scrollIntoView({{ block: 'center' }});
                radio.click();
            }}
        }}
        """
    )
    print("已點擊：自動代入申請人信息")

    page.evaluate(
        f"""
        () => {{
            const divs = Array.from(document.querySelectorAll('{container_selector}'));
            const visibleDiv = divs.find(d => d.offsetParent !== null && d.style.display !== 'none');
            if (!visibleDiv) return;

            const select = visibleDiv.querySelector('select[name="{sign_select_name}"]');
            if (select) {{
                select.value = '本名';
                select.dispatchEvent(new Event('change', {{ bubbles: true }}));
            }}
        }}
        """
    )
    print("已選擇署名情况：本名")

# ===============================
# Stages
# ===============================
def stage_0_guide(page):
    log("Stage 0", "導引 & 進入新建登記")

    page.goto(LOGIN_URL,timeout=10000)
    page.wait_for_url("**/user/welcome", timeout=10000)
    page.wait_for_load_state("networkidle")
    print("已進入:「登入後頁面」")

    page.goto(ADD_NEW_URL,timeout=10000, wait_until="networkidle")
    print("已進入:「新建登記頁」")

def stage_1_agreement(page):
    log("Stage 1", "同意條款")
    
    page.locator("#dczdj-input").wait_for(state="visible",timeout=5000)
    page.locator("#dczdj-input").check()
    print("已勾選:同意條款")

    click_item(page, "button", "开始办理")
    print("已點擊:开始办理")


def stage_2_1_product_info(page,WORK_NAME):
    log("Stage 2-1", "填寫作品信息")
   
    click_item(page, "input#name")
    page.locator("input#name").fill(WORK_NAME)
    print(f"已填寫:作品名稱: {WORK_NAME}")

    click_item(page, "button#productTypeBtn")
    print("已點擊:選擇類別")
  
    layer = page.locator("div.layui-layer:has(div.layui-layer-content)").first
    layer.wait_for(state="attached", timeout=10000)
    page.wait_for_timeout(500)  # 動畫稍微等一下
    page.locator("span#tree_10_check").click(force=True)
    print("已點擊:美術作品")

    page.locator("a.layui-layer-btn0:has-text('确定')").click()
    print("已點擊:確定")

    click_item(page, "button.dczdj-btn-2-1")
    print("已點擊:下一步")

def stage_2_2_product_info(page, WORK_DONE_DATE):
    log("Stage 2-2", "填寫作品信息")

    radio_original = page.locator("input[name='pro_nature'][value='0']:not([disabled])")
    radio_original.wait_for(state="visible", timeout=5000)
    radio_original.check()
    print("已選擇作品性質:原创")

    work_date_str = WORK_DONE_DATE.strftime("%Y-%m-%d")
    date_input = page.locator("#workTime")
    date_input.wait_for(state="visible", timeout=5000)
    # 移除 readonly
    page.evaluate("""
        document.querySelector('#workTime').removeAttribute('readonly')
    """)
    date_input.fill(work_date_str)
    close_datepicker(page)
    print(f"已填寫製作日期: {work_date_str}")

    country_select = page.locator("#WorkCountry")
    country_select.wait_for(state="visible", timeout=5000)
    country_select.select_option(label="中国")
    print("已选择国家:中国")

    province_select = page.locator("#WorkProvince")
    page.wait_for_function("""
        () => {
            const sel = document.querySelector('#WorkProvince');
            return sel && sel.options.length > 1;
        }
    """, timeout=5000)
    province_select.select_option(label="北京")
    print("已选择省份:北京")

    city_select = page.locator("#WorkCity")
    page.wait_for_function("""
        () => {
            const sel = document.querySelector('#WorkCity');
            return sel && sel.options.length >= 1;
        }
    """, timeout=5000)
    city_select.select_option(label="北京")
    print("已选择城市:北京")

    unpublished_radio = page.locator("input[name='pub_type'][value='0']").first
    unpublished_radio.wait_for(state="visible", timeout=5000)
    unpublished_radio.check()
    print("已選擇發表狀態:未发表")

    click_item(page, "button.dczdj-btn-2-2")
    print("已點擊:下一步")


def stage_2_3_product_info(page,DECLARATION_OWNERSHIP_PATH):
    log("Stage 2-3", "填寫作品信息")

    original_radio = page.get_by_label("原始")
    original_radio.wait_for(state="visible", timeout=5000)
    original_radio.check()
    print("已選擇权利取得方式:原始")

    click_item(page,"label#qlgsfs3")
    print("已選擇权利歸屬方式:法人作品")

    upload_file(page, ".beforeNamediv:has-text('法人作品权利归属声明')", DECLARATION_OWNERSHIP_PATH)
    page.wait_for_timeout(1500)
    print(f"已上传: {DECLARATION_OWNERSHIP_PATH} 完成")

    click_item(page, "button.dczdj-btn-2-3")
    print("已點擊:下一步")


def stage_2_4_author_info(page):
    log("Stage 2-4", "填寫作者信息")

    batch_select = page.locator("#authorNum")
    batch_select.wait_for(state="visible", timeout=10000)
    batch_select.select_option(value="1")
    print("已選擇:1 人")

    fill_person_block(
        page=page,
        container_selector='div.xjdj-table-box[name="authorDiv"]',
        radio_name="radio4",
        sign_select_name="authorSignType")

    wait_overlay_gone(page)
    click_item(page, "button.dczdj-btn-2","保存，下一步",500)
    print("已點擊:保存，下一步")

def stage_2_5_copyright_info(page):
    log("Stage 2-5", "填寫著作權人信息")

    batch_select = page.locator("#corSelect")
    batch_select.wait_for(state="visible", timeout=30000)
    batch_select.select_option(value="1")
    print("已選擇:1 人")

    fill_person_block(
        page=page,
        container_selector='div[name="corTempAddDiv"]',
        radio_name="radio41",
        sign_select_name="sign_type")

    click_item(page, "button.dczdj-btn-3","下一步")
    print("已點擊:下一步")

def stage_2_6_agent_info(page):
    log("Stage 2-6", "填寫申请人/代理人信息信息")

    wait_overlay_gone(page)
    click_item(page, "button.dczdj-btn-4","保存，下一步")
    print("已點擊:保存，下一步")


def stage_2_7_preview_info(page):
    log("Stage 2-7", "预览并提交作品登记申请表")

    wait_overlay_gone(page)
    click_item(page, "input.dczdj-btn[value='下一页']")
    print("已點擊:下一页")

    wait_overlay_gone(page)
    click_item(page, "button.dczdj-btn-5","保存，下一步")
    print("已點擊:保存，下一步")


def stage_2_8_product_description(page,PRODUCT_HIGHTLIGHT,PRODUCT_COPYRIGHT_HOLDER,PRODUCT_CREATIVE_PROCESS):
    log("Stage 2-8", "作品说明书")

    textarea_highlight = page.locator("textarea#central_content")
    textarea_highlight.wait_for(state="visible", timeout=10000)
    textarea_highlight.scroll_into_view_if_needed()
    textarea_highlight.fill(PRODUCT_HIGHTLIGHT)
    print("已填入:中心内容及作品特点")


    textarea_holder = page.locator("textarea#introduction")
    textarea_holder.wait_for(state="visible", timeout=10000)
    textarea_holder.scroll_into_view_if_needed()
    textarea_holder.fill(PRODUCT_COPYRIGHT_HOLDER)
    print("已填入:著作权人简介")

    textarea_process = page.locator("textarea#process")
    textarea_process.wait_for(state="visible", timeout=10000)
    textarea_process.scroll_into_view_if_needed()
    textarea_process.fill(PRODUCT_CREATIVE_PROCESS)
    print("已填入:创作过程")

    wait_overlay_gone(page)
    click_item(page, "button.dczdj-btn-6","下一步")
    print("已點擊:下一步")

def stage_2_9_preview_description(page):
    log("Stage 2-9", "预览作品說明書")

    wait_overlay_gone(page)
    page.wait_for_selector(".dczdj-model-7", state="visible", timeout=5000)
    click_item(page, ".dczdj-model-7 button.dczdj-btn","保存，下一步")
    page.wait_for_selector(".dczdj-model-8", state="visible", timeout=10000)
    print("已點擊:保存，下一步")

def stage_2_10_preview_waranty(page):
    log("Stage 2-10", "预览权利保证书")

    wait_overlay_gone(page)
    click_item(page, "button.dczdj-btn-8","保存，下一步")
    print("已點擊:保存，下一步")


def stage_2_11_upload_product(page,PRODUCT_DOCUMENT_PATH):
    log("Stage 2-11", "传作品文件")

    upload_file(page, ".beforeNamediv:has-text('作品文件')", PRODUCT_DOCUMENT_PATH)
    page.wait_for_timeout(1500)
    print(f"已上传： {PRODUCT_DOCUMENT_PATH} 完成")

    wait_overlay_gone(page)
    click_item(page, "button.dczdj-btn-9","保存，下一步",500)
    print("已點擊:保存，下一步")

def stage_2_12_download_file(page,WORK_NAME,DOCUMENT_DOWNLOAD_PATH):
    log("Stage 2-12", "确认申报确认函，并上传签字拍照文件")

    # 選擇所有浏览确认檔案
    rows = page.locator("#affirmShow .dczdj-8-container-detail").filter(
        has=page.locator("a", has_text="浏览确认")
    )
    rows.first.wait_for(state="visible", timeout=10000)
    count = rows.count()
    print(f"浏览确认檔案數量：{count}")

    for i in range(count):
        row = rows.nth(i)
        link = row.locator("a", has_text="浏览确认")
        title_div = row.locator("div").first
        raw_title = title_div.inner_text().strip()
        checkbox = row.locator("input.exploreCheck")

        # 1️⃣ 取得 URL
        onclick_attr = link.get_attribute("onclick")  # e.g. showFile(this,'http://file.bjxwcbj.gov.cn/...pdf')
        file_url = onclick_attr.split("'")[1]  # 取單引號中間的 URL
        file_name_from_url = os.path.basename(urlparse(file_url).path)  # 取檔名
        # 自動命名：序號-原檔名
        save_path = os.path.join(DOCUMENT_DOWNLOAD_PATH, f"{WORK_NAME}-{raw_title}-{file_name_from_url}")

        # 2️⃣ 下載檔案
        with requests.get(file_url, stream=True) as r:
            r.raise_for_status()
            with open(save_path, "wb") as f:
                for chunk in r.iter_content(8192):
                    f.write(chunk)
        print(f"已下載:{save_path}")

        prevent_pdf_new_tab(page)
        link.click()

        print(f"✔ 已浏览确认第 {i+1} 個檔案")

    wait_overlay_gone(page)
    click_item(page, "button.dczdj-btn-11-1","下一步",500)
    print("已點擊:保存，下一步")


def stage_2_13_download_confirm_file(page):
    log("Stage 2-13", "下載打印申報確認函")

    page.evaluate("() => { setTimeout(() => onprinta(), 100); }")

    log("[END]", "請在瀏覽器中完成列印或取消")

    # 👇 等使用者按下「繼續」
    app_context.ui.wait_for_choice(
        "請在瀏覽器中完成列印或取消\n\n完成後請按【繼續】進行下一筆測試",
        primary_text="▶【繼續】"
    )

# ===============================
# Workflow
# ===============================
def run_workflow(storage_state_path: str, excel_path: str):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    # 欄位名稱 → index 對應
    header = [cell.value for cell in ws[1]]

    REQUIRED_COLUMNS = [
        "作品名稱",
        "完成日期(yyyy-MM-dd)",
        "權利歸屬聲明絕對路徑",
        "作品說明(中心內容及特點)",
        "作品說明(著作權人簡介)",
        "作品說明(創作過程)",
        "作品文件絕對路徑",
        "文件下載路徑"
    ]
    missing = set(REQUIRED_COLUMNS) - set(header)
    if missing:
        raise ValueError(f"Excel 缺少欄位：{missing}")

    col_index = {name: header.index(name) for name in REQUIRED_COLUMNS}

    chrome_path = get_system_chrome_path()

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            executable_path=chrome_path,
            args=["--start-maximized"]
        )

        context = browser.new_context(
            storage_state=storage_state_path,
            no_viewport=True
        )

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            try:
                WORK_NAME = row[col_index["作品名稱"]]
                # ⛔ 空列直接跳過（這一行是關鍵）
                if WORK_NAME is None:
                    continue
                WORK_DONE_DATE = row[col_index["完成日期(yyyy-MM-dd)"]]
                if isinstance(WORK_DONE_DATE, datetime):
                    WORK_DONE_DATE = WORK_DONE_DATE.date()

                DECLARATION_OWNERSHIP_PATH = row[col_index["權利歸屬聲明絕對路徑"]]
                PRODUCT_HIGHTLIGHT = row[col_index["作品說明(中心內容及特點)"]]
                PRODUCT_COPYRIGHT_HOLDER = row[col_index["作品說明(著作權人簡介)"]]
                PRODUCT_CREATIVE_PROCESS = row[col_index["作品說明(創作過程)"]]
                PRODUCT_DOCUMENT_PATH = row[col_index["作品文件絕對路徑"]]
                DOCUMENT_DOWNLOAD_PATH = row[col_index["文件下載路徑"]]


                log("程式流程開始",f"\n=== 處理第 {idx} 筆 ===\n\n作品: {WORK_NAME} 日期: {WORK_DONE_DATE}\n權利歸屬聲明(路徑)：{DECLARATION_OWNERSHIP_PATH}\n作品文件(路徑)：{PRODUCT_DOCUMENT_PATH} \n文件下載(路徑)：{DOCUMENT_DOWNLOAD_PATH}\n\n作品說明(中心內容及特點)：\n{PRODUCT_HIGHTLIGHT}\n\n作品說明(著作權人簡介)：\n{PRODUCT_COPYRIGHT_HOLDER}\n\n作品說明(創作過程)：\n{PRODUCT_CREATIVE_PROCESS}\n")
            
                page = context.new_page()

                stage_0_guide(page)
                stage_1_agreement(page)
                stage_2_1_product_info(page,WORK_NAME)
                stage_2_2_product_info(page,WORK_DONE_DATE)
                stage_2_3_product_info(page,DECLARATION_OWNERSHIP_PATH)
                stage_2_4_author_info(page)
                stage_2_5_copyright_info(page)
                stage_2_6_agent_info(page)
                stage_2_7_preview_info(page)
                stage_2_8_product_description(page,PRODUCT_HIGHTLIGHT,PRODUCT_COPYRIGHT_HOLDER,PRODUCT_CREATIVE_PROCESS)
                stage_2_9_preview_description(page)
                stage_2_10_preview_waranty(page)
                stage_2_11_upload_product(page,PRODUCT_DOCUMENT_PATH)
                stage_2_12_download_file(page,WORK_NAME,DOCUMENT_DOWNLOAD_PATH)
                stage_2_13_download_confirm_file(page)

                page.close()
            except Exception as e:
                text = f"第 {idx+1} 筆資料發生錯誤: {e}"
                log("ERROR", text)
                page.close()

        browser.close()
